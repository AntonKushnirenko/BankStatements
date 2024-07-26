# Библиотеки для работы с excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.utils.cell import get_column_letter

from math import floor

from sort_xlsx_by_column import sort_file_by_column

# Столбцы
tpd_column = "A"  # Столбец ТПД
client_name_column = "C"  # Столбец Название клиента
sku_column = "E"  # Столбец SKU
sales_volume_column = "F"  # Столбец Объем продаж
SKU_of_MML_column = "G"  # Столбец SKU по MML
sales_volume_from_MML_column = "H"  # Столбец Объем продаж из MML
point_counted_tpd_column = "I"  # Столбец Точка засчитана ТПД
point_counted_rtt_column = "J"  # Столбец Точка засчитана РТТ
number_of_cards_column = "K"  # Столбец Количество карт

def get_input_file_name():
    input_file_name = input("Enter file name: ")
    if not input_file_name.endswith(".xlsx"):
        add_xlsx_answer = input("Not an excel file. Add .xlsx? Y/N ").upper()
        if add_xlsx_answer == "Y" or add_xlsx_answer == "YES":
            input_file_name += ".xlsx"
        else:
            return None
    return input_file_name
def reset_input_file_name():
    input_file_name = None
    while not input_file_name:
        input_file_name = get_input_file_name()
    return input_file_name

def sort_file(input_file_name):
    return sort_file_by_column(input_file_name, "Отчет по ТПД", "ТПД")

input_file_name = reset_input_file_name()
sorted_input_file_name = sort_file(input_file_name)
while not sorted_input_file_name:
    input_file_name = reset_input_file_name()
    sorted_input_file_name = sort_file(input_file_name)
print("Valid name. Processing file")

fill_subtotal_row = PatternFill(fill_type="solid", fgColor='00AADCF7')

# Тонкие границы
thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000'))

# Посчитать количество не пустых строк в столбце в определенном диапазоне
def count_non_empty_rows(sheet, column, first_row, last_row):
    count = 0
    for row_number in range(first_row, last_row+1):
        if sheet[f"{column}{row_number}"].value:
            count += 1
    return count

# Собираем строки в группы и пишем промежуточный итог
def group_and_write_subtotal(sheet):
    # Значение ТПД для сравнения
    TPD_cell_to_compare = sheet["A2"]
    TPD_group_first_row = TPD_cell_to_compare.row
    for cell in sheet[tpd_column][1:]:
        # Проверяем одинаковый ли ТПД
        if cell.value == TPD_cell_to_compare.value:
            TPD_group_last_row = cell.row
        else:
            group_and_write_subtotal_helper(sheet, TPD_cell_to_compare, TPD_group_first_row, TPD_group_last_row, 1)
            TPD_cell_to_compare = cell
            TPD_group_first_row = cell.row
    group_and_write_subtotal_helper(sheet, TPD_cell_to_compare, TPD_group_first_row, TPD_group_last_row, 1)

def group_and_write_subtotal_helper(sheet, TPD_cell_to_compare, TPD_group_first_row, TPD_group_last_row, outline_level):
    if TPD_group_first_row < TPD_group_last_row:  # Если в группе больше одной строки
        # Группируем строки перед обновления значения ТПД на новое
        sheet.row_dimensions.group(TPD_group_first_row, TPD_group_last_row, outline_level=outline_level, hidden=True)
    else:
        TPD_group_last_row = TPD_group_first_row

    subtotal_row = TPD_group_last_row + 1
    sheet.insert_rows(subtotal_row, amount=1)  # Вставляем строку между группами для результата
    sheet[f"{tpd_column}{subtotal_row}"].value = f"{TPD_cell_to_compare.value} Результат "
    write_subtotal_values(subtotal_row, TPD_group_first_row, TPD_group_last_row)
    fill_cells_in_row(subtotal_row)

def write_subtotal_values(row_number, first_row, last_row):
    #sheet[f"C{row_number}"].value = last_row - first_row + 1
    sheet[f"{client_name_column}{row_number}"].value = f"=COUNT({sku_column}{first_row}:{sku_column}{last_row})"
    sheet[f"{sales_volume_column}{row_number}"].value = f"=SUBTOTAL(9,{sales_volume_column}{first_row}:{sales_volume_column}{last_row})"
    #sheet[f"{SKU_of_MML_column}{row_number}"].value = count_non_empty_rows(sheet, SKU_of_MML_column, first_row, last_row)
    sheet[f"{SKU_of_MML_column}{row_number}"].value = f"=COUNTA({SKU_of_MML_column}{first_row}:{SKU_of_MML_column}{last_row})-COUNTIF({SKU_of_MML_column}{first_row}:{SKU_of_MML_column}{last_row}, 0)"
    sheet[f"{sales_volume_from_MML_column}{row_number}"].value = f"=SUBTOTAL(9,{sales_volume_from_MML_column}{first_row}:{sales_volume_from_MML_column}{last_row})"
    sheet[f"{point_counted_tpd_column}{row_number}"].value = f"=SUBTOTAL(9,{point_counted_tpd_column}{first_row}:{point_counted_tpd_column}{last_row})"
    sheet[f"{point_counted_rtt_column}{row_number}"].value = f"=SUBTOTAL(9,{point_counted_rtt_column}{first_row}:{point_counted_rtt_column}{last_row})"
    sheet[f"{number_of_cards_column}{row_number}"].value = f"=SUBTOTAL(9,{number_of_cards_column}{first_row}:{number_of_cards_column}{last_row})"

def fill_cells_in_row(row_number):
    for cell in sheet[row_number]:
        cell.fill = fill_subtotal_row

def group_total(sheet):
    sheet.row_dimensions.group(2, sheet.max_row, outline_level=0, hidden=False)

# Общий итог, самая последняя строчка
def write_total(sheet):
    first_row = 2
    last_row = sheet.max_row
    total_row = last_row + 1

    sheet[f"{tpd_column}{total_row}"].value = "Общий итог Сумма"
    write_subtotal_values(total_row, first_row, last_row)
    # Вычитаем промежуточные строки с результатом для общего итога
    sheet[f"{SKU_of_MML_column}{total_row}"].value += f'-COUNTIF({tpd_column}{first_row}:{tpd_column}{last_row}, "*Результат*")'

    fill_cells_in_row(total_row)

def format_columns_into_rubles(sheet, column_letters):
    for row_number in range(2, sheet.max_row + 1):
        for column_letter in column_letters:
            sheet[f"{column_letter}{row_number}"].number_format = '#,##0₽'


# Правило для мотивации ТПД и РТТ
def write_motivation_rule(sheet, rule_name,  min_SKU_of_MML_amount, min_sales_volume_amount, column):
    # Подписываем колонку с правилом
    sheet[f"{column}1"].value = rule_name
    style_title_cell(sheet[f"{column}1"])
    sheet.column_dimensions[column].width = 20

    # Получение значений из строчек
    for row_number in range(2, sheet.max_row + 1):
        SKU_of_MML_value = sheet[f"{SKU_of_MML_column}{row_number}"].value
        sales_volume_value = sheet[f"{sales_volume_column}{row_number}"].value
        if SKU_of_MML_value:
            if sales_volume_value:
                SKU_of_MML_value = int(SKU_of_MML_value)
                sales_volume_value = float(sales_volume_value)
                # Проверка выполнения условий правила
                if SKU_of_MML_value >= min_SKU_of_MML_amount and sales_volume_value >= min_sales_volume_amount:
                    sheet[f"{column}{row_number}"].value = 1
                else:
                    sheet[f"{column}{row_number}"].value = 0
            else:
                sheet[f"{sales_volume_column}{row_number}"].value = 0
                sheet[f"{column}{row_number}"].value = 0
        else:
            sheet[f"{SKU_of_MML_column}{row_number}"].value = 0
            sheet[f"{sales_volume_from_MML_column}{row_number}"].value = 0
            sheet[f"{column}{row_number}"].value = 0

def write_number_of_cards(sheet):
    min_sales_volume_column = 5000
    sheet[f"{number_of_cards_column}1"].value = "Количество карт"
    style_title_cell(sheet[f"{number_of_cards_column}1"])
    sheet.column_dimensions[number_of_cards_column].width = 20
    for row_number in range(2, sheet.max_row + 1):
        sheet[f"{number_of_cards_column}{row_number}"].value = floor(sheet[f"{sales_volume_column}{row_number}"].value / min_sales_volume_column) * sheet[f"{point_counted_rtt_column}{row_number}"].value
        #sheet[f"{number_of_cards_column}{row_number}"].value = f"=ROUNDDOWN({sales_volume_column}{row_number}/{min_sales_volume_column},0)*{point_counted_rtt_column}{row_number}"

# Стилизация
def style_title_cell(cell):
    cell.font = Font(bold=True, name='Calibri')
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

def set_columns_width(sheet, width):
    max_column = 0
    for cell in sheet[1]:
        if cell.value != "":
            max_column = cell.column
    for column in range(1, max_column+1):
        sheet.column_dimensions[get_column_letter(column)].width = width


workbook = load_workbook(sorted_input_file_name)
sheet = workbook.active

set_columns_width(sheet, 25)

# Правило для мотивации ТПД
# Отгрузка 5 SKU из MML
write_motivation_rule(sheet, "Точка засчитана ТПД",5, 0, point_counted_tpd_column)

# Правило для мотивации РТТ
# Отгрузка 10 SKU из MML
# Объем продаж не менее 5000 рублей
write_motivation_rule(sheet, "Точка засчитана РТТ",10, 5000, point_counted_rtt_column)

# Записать столбец количество карт
write_number_of_cards(sheet)

# Сгруппировать и написать промежуточные результаты
group_total(sheet)  # 0 уровень (общий) должен быть до 1, из-за того, что так работает библиотека
group_and_write_subtotal(sheet)
write_total(sheet)

# Форматируем столбцы с объемами продаж в рубли
format_columns_into_rubles(sheet, [sales_volume_column, sales_volume_from_MML_column])

#workbook.save("subtotal1.xlsx")
workbook.save(input_file_name)
